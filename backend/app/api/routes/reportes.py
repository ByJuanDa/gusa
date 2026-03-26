from io import BytesIO
from datetime import datetime, date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from app.db.database import get_db
from app.models.llanta import Llanta
from app.models.venta import Venta, DetalleVenta
from app.services.auth_service import require_admin, require_reporte
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])

# ── Lista de precios pública ──────────────────────────────────

@router.get("/lista-precios")
def lista_precios_publica(db: Session = Depends(get_db)):
    """
    XLSX público con la lista de precios de llantas en catálogo.
    Sin autenticación requerida.
    """
    llantas = (
        db.query(Llanta)
        .options(joinedload(Llanta.marca))
        .filter(Llanta.visible_catalogo == True, Llanta.stock > 0)
        .order_by(Llanta.id_marca, Llanta.medida)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"
    ws.sheet_view.showGridLines = False

    # Título principal
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "GUSA — Lista de Precios"
    t.font = Font(bold=True, color=NEGRO, size=14)
    t.fill = PatternFill("solid", fgColor=AMARILLO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Fecha
    ws.merge_cells("A2:E2")
    f = ws["A2"]
    f.value = f"Vigente al {datetime.now().strftime('%d de %B de %Y').lower()}"
    f.font = Font(italic=True, color="FF6B7280", size=9)
    f.fill = PatternFill("solid", fgColor=NEGRO)
    f.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Encabezados
    headers = ["Marca", "Modelo", "Medida", "Descripción", "Precio de venta"]
    for col, h in enumerate(headers, 1):
        _hdr_cell(ws, 3, col, h)
    ws.row_dimensions[3].height = 22

    # Agrupar por marca para mejor legibilidad
    marca_actual = None
    fila = 4
    for l in llantas:
        marca_nombre = (l.marca.nombre_display or l.marca.nombre) if l.marca else ""

        # Separador de marca
        if marca_nombre != marca_actual:
            if marca_actual is not None:
                fila += 1  # espacio entre marcas
            ws.merge_cells(f"A{fila}:E{fila}")
            mc = ws[f"A{fila}"]
            mc.value = f"  {marca_nombre.upper()}"
            mc.font = Font(bold=True, color=AMARILLO, size=10)
            mc.fill = PatternFill("solid", fgColor="FF0F1117")
            mc.alignment = Alignment(vertical="center")
            thin = Side(style="thin", color="FF1F2937")
            mc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.row_dimensions[fila].height = 20
            fila += 1
            marca_actual = marca_nombre

        bg = "FF0A0E17" if fila % 2 == 0 else "FF06080F"
        _data_cell(ws, fila, 1, marca_nombre,       bg=bg, fg="FF9CA3AF")
        _data_cell(ws, fila, 2, l.modelo or "—",    bg=bg, fg="FFE5E7EB", bold=True)
        _data_cell(ws, fila, 3, l.medida,            bg=bg, fg="FFFACC15", bold=True)
        _data_cell(ws, fila, 4, l.descripcion or "", bg=bg, fg="FF6B7280")
        c = _data_cell(ws, fila, 5, l.precio_venta or 0, bg=bg, fg="FF4ADE80", bold=True)
        c.number_format = '"$"#,##0.00'
        fila += 1

    # Nota al pie
    fila += 1
    ws.merge_cells(f"A{fila}:E{fila}")
    nota = ws[f"A{fila}"]
    nota.value = "Precios sujetos a cambio sin previo aviso. Consulta disponibilidad."
    nota.font = Font(italic=True, color="FF374151", size=9)
    nota.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A4"
    _autofit(ws)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"gusa_lista_precios_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# ── Helpers de estilo ─────────────────────────────────────────

AMARILLO  = "FFFACC15"
NEGRO     = "FF030712"
GRIS_OSC  = "FF111827"
GRIS_MED  = "FF1F2937"
BLANCO    = "FFFFFFFF"
VERDE     = "FF4ADE80"
ROJO      = "FFF87171"

def _hdr_cell(ws, row, col, value, bold=True, bg=AMARILLO, fg=NEGRO):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FF374151")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def _data_cell(ws, row, col, value, bold=False, number_format=None, fg=BLANCO, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=9)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="center")
    if number_format:
        c.number_format = number_format
    return c

def _autofit(ws, min_w=10, max_w=50):
    for col in ws.columns:
        length = max_w
        for cell in col:
            try:
                length = max(min_w, min(max_w, len(str(cell.value or "")) + 4))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = length


# ── Reporte de Inventario ─────────────────────────────────────

@router.get("/inventario")
def reporte_inventario(
    db: Session = Depends(get_db),
    usuario=Depends(require_reporte),
):
    """
    XLSX con el inventario completo.
    Acceso: Sistemas, Gerente General, Encargado.
    """
    llantas = (
        db.query(Llanta)
        .options(joinedload(Llanta.marca))
        .order_by(Llanta.codigo)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"GUSA — Reporte de Inventario  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    t.font = Font(bold=True, color=BLANCO, size=12)
    t.fill = PatternFill("solid", fgColor=NEGRO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Encabezados
    headers = ["#", "Código", "Marca", "Modelo", "Medida",
               "Stock", "Costo compra $", "Precio venta $", "Valor en almacén $", "Catálogo"]
    for col, h in enumerate(headers, 1):
        _hdr_cell(ws, 2, col, h)
    ws.row_dimensions[2].height = 22

    # Datos
    total_piezas = 0
    valor_total  = 0.0
    for i, l in enumerate(llantas, 1):
        row = i + 2
        bg = "FF0A0E17" if i % 2 == 0 else "FF06080F"
        marca = l.marca.nombre_display or l.marca.nombre if l.marca else ""
        valor = (l.costo_compra or 0) * (l.stock or 0)
        total_piezas += l.stock or 0
        valor_total  += valor

        _data_cell(ws, row, 1,  i,          bg=bg, fg="FF6B7280")
        _data_cell(ws, row, 2,  l.codigo,   bg=bg, fg="FFFACC15", bold=True)
        _data_cell(ws, row, 3,  marca,      bg=bg, fg="FFE5E7EB", bold=True)
        _data_cell(ws, row, 4,  l.modelo or "", bg=bg, fg="FF9CA3AF")
        _data_cell(ws, row, 5,  l.medida,   bg=bg, fg="FF9CA3AF")
        stock_color = ROJO if (l.stock or 0) == 0 else ("FFFFA040" if (l.stock or 0) <= 3 else BLANCO)
        _data_cell(ws, row, 6,  l.stock or 0, bg=bg, fg=stock_color, bold=True)
        _data_cell(ws, row, 7,  l.costo_compra or 0, bg=bg, number_format='"$"#,##0.00')
        _data_cell(ws, row, 8,  l.precio_venta or 0, bg=bg, fg="FF4ADE80", number_format='"$"#,##0.00')
        _data_cell(ws, row, 9,  valor,       bg=bg, number_format='"$"#,##0.00')
        _data_cell(ws, row, 10, "Sí" if l.visible_catalogo else "No", bg=bg, fg="FF6B7280")

    # Fila de totales
    tot_row = len(llantas) + 3
    ws.merge_cells(f"A{tot_row}:E{tot_row}")
    _hdr_cell(ws, tot_row, 1, "TOTALES", bg=GRIS_OSC, fg="FFFACC15")
    _hdr_cell(ws, tot_row, 6, total_piezas, bg=GRIS_OSC, fg=VERDE)
    _hdr_cell(ws, tot_row, 7, "", bg=GRIS_OSC, fg=BLANCO)
    _hdr_cell(ws, tot_row, 8, "", bg=GRIS_OSC, fg=BLANCO)
    c = _hdr_cell(ws, tot_row, 9, valor_total, bg=GRIS_OSC, fg=VERDE)
    c.number_format = '"$"#,##0.00'
    _hdr_cell(ws, tot_row, 10, "", bg=GRIS_OSC, fg=BLANCO)

    ws.freeze_panes = "A3"
    _autofit(ws)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"inventario_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Reporte de Ventas ─────────────────────────────────────────

@router.get("/ventas")
def reporte_ventas(
    fecha_desde: date | None = Query(None),
    fecha_hasta: date | None = Query(None),
    db: Session = Depends(get_db),
    usuario=Depends(require_admin),
):
    """
    XLSX con ventas en el rango de fechas indicado.
    Acceso: Sistemas y Gerente General.
    """
    q = db.query(Venta).options(
        joinedload(Venta.usuario),
        joinedload(Venta.detalles).joinedload(DetalleVenta.llanta).joinedload(Llanta.marca),
    )
    if fecha_desde:
        q = q.filter(Venta.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        q = q.filter(Venta.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))
    ventas = q.order_by(Venta.fecha.desc()).all()

    wb = openpyxl.Workbook()

    # ── Hoja 1: Ventas con detalle inline ─────────────────────
    ws1 = wb.active
    ws1.title = "Ventas"
    ws1.sheet_view.showGridLines = False

    rango = ""
    if fecha_desde and fecha_hasta:
        rango = f"  |  {fecha_desde.strftime('%d/%m/%Y')} — {fecha_hasta.strftime('%d/%m/%Y')}"
    elif fecha_desde:
        rango = f"  |  Desde {fecha_desde.strftime('%d/%m/%Y')}"
    elif fecha_hasta:
        rango = f"  |  Hasta {fecha_hasta.strftime('%d/%m/%Y')}"

    NUM_COLS = 12
    ws1.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    t = ws1["A1"]
    t.value = f"GUSA — Reporte de Ventas{rango}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    t.font = Font(bold=True, color=BLANCO, size=12)
    t.fill = PatternFill("solid", fgColor=NEGRO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 26

    # Columnas: A=# B=Folio C=Fecha D=Vendedor E=Código F=Marca G=Modelo H=Medida I=Cant. J=P.Unit.$ K=Total/Sub$ L=Estado
    hdrs1 = ["#", "Folio", "Fecha", "Vendedor", "Código", "Marca", "Modelo", "Medida", "Cant.", "P. Unit. $", "Total / Sub $", "Estado"]
    for col, h in enumerate(hdrs1, 1):
        _hdr_cell(ws1, 2, col, h)
    ws1.row_dimensions[2].height = 22

    ventas_activas = 0
    total_ingresos = 0.0
    total_piezas_g = 0

    BG_VENTA_PAR  = "FF0D1220"
    BG_VENTA_IMPAR = "FF080C18"
    BG_PROD       = "FF060A14"

    cur_row = 3
    for i, v in enumerate(ventas, 1):
        piezas = sum(d.cantidad for d in v.detalles)
        vendedor = v.usuario.nombre if v.usuario else "—"
        estado_color = VERDE if v.status == "activa" else ROJO
        bg_v = BG_VENTA_PAR if i % 2 == 0 else BG_VENTA_IMPAR

        if v.status == "activa":
            ventas_activas += 1
            total_ingresos += v.total or 0
            total_piezas_g += piezas

        # ── fila cabecera de la venta ──────────────────────────
        _data_cell(ws1, cur_row, 1,  i,                       bg=bg_v, fg="FF6B7280")
        _data_cell(ws1, cur_row, 2,  f"V-{v.id:04d}",         bg=bg_v, fg="FFFACC15", bold=True)
        _data_cell(ws1, cur_row, 3,  v.fecha.strftime("%d/%m/%Y %H:%M") if v.fecha else "", bg=bg_v, fg="FF9CA3AF")
        _data_cell(ws1, cur_row, 4,  vendedor,                 bg=bg_v, fg="FFE5E7EB", bold=True)
        # Columnas E-J vacías en la fila de venta (el detalle va abajo)
        for col in range(5, 11):
            _data_cell(ws1, cur_row, col, "", bg=bg_v)
        c = _data_cell(ws1, cur_row, 11, v.total or 0,        bg=bg_v, fg=VERDE, bold=True)
        c.number_format = '"$"#,##0.00'
        _data_cell(ws1, cur_row, 12, v.status.upper(),        bg=bg_v, fg=estado_color, bold=True)
        ws1.row_dimensions[cur_row].height = 18
        cur_row += 1

        # ── filas de detalle (productos) ───────────────────────
        for d in v.detalles:
            ll = d.llanta
            marca = (ll.marca.nombre_display or ll.marca.nombre) if ll and ll.marca else "—"
            for col in range(1, 5):
                _data_cell(ws1, cur_row, col, "", bg=BG_PROD)
            _data_cell(ws1, cur_row, 5,  ll.codigo if ll else "—",   bg=BG_PROD, fg="FFFACC15")
            _data_cell(ws1, cur_row, 6,  marca,                       bg=BG_PROD, fg="FFD1D5DB")
            _data_cell(ws1, cur_row, 7,  ll.modelo or "" if ll else "", bg=BG_PROD, fg="FF9CA3AF")
            _data_cell(ws1, cur_row, 8,  ll.medida if ll else "",     bg=BG_PROD, fg="FFFBBF24")
            _data_cell(ws1, cur_row, 9,  d.cantidad,                  bg=BG_PROD, fg=BLANCO, bold=True)
            c1 = _data_cell(ws1, cur_row, 10, d.precio_unitario,      bg=BG_PROD, fg="FF86EFAC")
            c1.number_format = '"$"#,##0.00'
            c2 = _data_cell(ws1, cur_row, 11, d.subtotal,             bg=BG_PROD, fg="FF4ADE80", bold=True)
            c2.number_format = '"$"#,##0.00'
            _data_cell(ws1, cur_row, 12, v.notas or "",               bg=BG_PROD, fg="FF6B7280")
            ws1.row_dimensions[cur_row].height = 16
            cur_row += 1

        # espacio entre ventas
        cur_row += 1

    # Fila de totales
    tot_row = cur_row
    ws1.merge_cells(f"A{tot_row}:D{tot_row}")
    _hdr_cell(ws1, tot_row, 1, f"TOTALES  ({ventas_activas} ventas activas)", bg=GRIS_OSC, fg="FFFACC15")
    for col in range(5, 10):
        _hdr_cell(ws1, tot_row, col, "", bg=GRIS_OSC)
    _hdr_cell(ws1, tot_row, 9, total_piezas_g, bg=GRIS_OSC, fg=VERDE)
    c = _hdr_cell(ws1, tot_row, 11, total_ingresos, bg=GRIS_OSC, fg=VERDE)
    c.number_format = '"$"#,##0.00'
    _hdr_cell(ws1, tot_row, 12, "", bg=GRIS_OSC)

    ws1.freeze_panes = "A3"
    # Anchos fijos para legibilidad
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 16
    ws1.column_dimensions["G"].width = 16
    ws1.column_dimensions["H"].width = 12
    ws1.column_dimensions["I"].width = 7
    ws1.column_dimensions["J"].width = 14
    ws1.column_dimensions["K"].width = 14
    ws1.column_dimensions["L"].width = 14

    # ── Hoja 2: Detalle por línea de producto ─────────────────
    ws2 = wb.create_sheet("Detalle")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:J1")
    t2 = ws2["A1"]
    t2.value = "GUSA — Detalle de líneas vendidas"
    t2.font = Font(bold=True, color=BLANCO, size=12)
    t2.fill = PatternFill("solid", fgColor=NEGRO)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    hdrs2 = ["Folio", "Fecha", "Vendedor", "Código", "Marca", "Modelo", "Medida",
             "Cantidad", "Precio unit. $", "Subtotal $"]
    for col, h in enumerate(hdrs2, 1):
        _hdr_cell(ws2, 2, col, h)
    ws2.row_dimensions[2].height = 22

    row2 = 3
    for v in ventas:
        if v.status == "cancelada":
            continue
        vendedor = v.usuario.nombre if v.usuario else "—"
        fecha_str = v.fecha.strftime("%d/%m/%Y %H:%M") if v.fecha else ""
        for d in v.detalles:
            l = d.llanta
            bg = "FF0A0E17" if row2 % 2 == 0 else "FF06080F"
            marca = (l.marca.nombre_display or l.marca.nombre) if l and l.marca else ""
            _data_cell(ws2, row2, 1,  f"V-{v.id:04d}", bg=bg, fg="FFFACC15", bold=True)
            _data_cell(ws2, row2, 2,  fecha_str,         bg=bg, fg="FF9CA3AF")
            _data_cell(ws2, row2, 3,  vendedor,          bg=bg, fg="FFE5E7EB")
            _data_cell(ws2, row2, 4,  l.codigo if l else "—", bg=bg, fg="FFFACC15")
            _data_cell(ws2, row2, 5,  marca,             bg=bg, fg="FFE5E7EB")
            _data_cell(ws2, row2, 6,  l.modelo or "" if l else "", bg=bg, fg="FF9CA3AF")
            _data_cell(ws2, row2, 7,  l.medida if l else "", bg=bg, fg="FF9CA3AF")
            _data_cell(ws2, row2, 8,  d.cantidad,        bg=bg, fg=BLANCO, bold=True)
            c1 = _data_cell(ws2, row2, 9, d.precio_unitario, bg=bg, fg="FF4ADE80")
            c1.number_format = '"$"#,##0.00'
            c2 = _data_cell(ws2, row2, 10, d.subtotal,  bg=bg, fg="FF4ADE80", bold=True)
            c2.number_format = '"$"#,##0.00'
            row2 += 1

    ws2.freeze_panes = "A3"
    _autofit(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ventas_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
